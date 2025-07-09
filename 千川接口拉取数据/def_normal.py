#常用方法集合
from json import dumps
from urllib.parse import quote as parse_quote
import os
from time import sleep
from itertools import zip_longest

from requests import post,get
import openpyxl
import pandas as pd

#json整合url参数,用于整理请求url

def process_value(v):
    if isinstance(v, dict) or isinstance(v, list):
        return dumps(v, ensure_ascii=False)
    else:
        return str(v)

def build_url(base_url, params):
    # 列出需要JSON编码的参数名
    json_params = ['dimensions', 'metrics', 'filtering', 'fields', 'order_by', 'filters']
    
    # 预处理filters内的values
    if 'filters' in params:
        for filter_dict in params['filters']:
            if 'values' in filter_dict:
                filter_dict['values'] = [process_value(v) for v in filter_dict['values']]
                
    for param in json_params:
        if param in params:
            # 使用dumps将参数值转换为JSON字符串，并确保非ASCII字符也被正确编码
            params[param] = process_value(params[param])
    
    encoded_params = {key: parse_quote(process_value(value)) for key, value in params.items()}
    query_string = '&'.join([f'{key}={value}' for key, value in encoded_params.items()])
    
    return f'{base_url}?{query_string}'

#读取本地请求密钥，用于获取请求token
def get_token():
    dir = os.path.dirname(__file__)
    filepath = os.path.join(dir, '最新刷新码.xlsx')
    df = pd.read_excel(filepath, sheet_name=0, usecols='B:C', nrows=4, header=None,dtype='str')
    tokens_dict = {
        'plan': {'refresh_token': df.iloc[1, 0],'APP_ID': df.iloc[2, 0],'secret': df.iloc[3, 0]},
        'aweme': {'refresh_token': df.iloc[1, 1],'APP_ID': df.iloc[2, 1],'secret': df.iloc[3, 1]}
    }
    return tokens_dict

#将最新的refresh_token更新到本地
def re_refresh_token(refresh_token,token_type):
    dir = os.path.dirname(__file__)
    filepath = os.path.join(dir, '最新刷新码.xlsx')
    column_index = {'plan': 2,'aweme': 3}.get(token_type,2)
    df = pd.read_excel(filepath, sheet_name=0, header=None)
    df.iloc[1, column_index-1] = refresh_token
    df.to_excel(filepath, index=False, header=False)

#刷新access_token
def re_access_token(APP_ID,secret,refresh_token,retry_limit=3):
    open_api_url = 'https://ad.oceanengine.com/open_api/'
    url_params = 'oauth2/refresh_token/'
    url = open_api_url + url_params
    json_params = {
        'appid': APP_ID,
        'secret': secret,
        'grant_type': 'refresh_token',
        'refresh_token': refresh_token,
    }
    headers = {'Content-Type': 'application/json'}
    try:
        retry_count = 1
        while retry_count <= retry_limit:
            format_url = build_url(url,json_params)
            rsp = post(format_url, headers=headers)
            rsp_data = rsp.json()
            if rsp_data.get('code') == 0:
                access_token = rsp_data['data'].get('access_token',None)
                new_refresh_token = rsp_data['data'].get('refresh_token',None)
                if access_token and new_refresh_token:
                    return access_token, new_refresh_token
                else:
                    retry_count += 1
                    sleep(2)
            else:
                retry_count += 1
                sleep(2)
        else:
            print(f're_access_token请求多次失败: {rsp_data}')
            return None
    except Exception as e:
        print(f're_access_token未成功发起请求: {e}')
        return None
    
#创建excel表格
def create_plan_excel(yesterday):
    dir = os.path.dirname(__file__)
    filepath = os.path.join(dir,'源表','推广费',f'推广费-{yesterday}.xlsx')
    if os.path.exists(filepath):
        print('文件已存在！即将进行计划数据获取......')
    else:
        print('创建计划文件中......')
        table_title = ['推广类型', '账户名称', '账户ID', '计划ID', '计划名称', '商品ID', '商品名称', '抖音号ID', '抖音号名称', '日期', '消耗(元)', '展示次数', '点击率(%)', '点击次数','转化数', '转化率(%)', '转化成本(元)', '7日总支付ROI', '直接成交订单数','直接成交金额(元)', '直接支付ROI', '直接下单订单数', '直接下单金额(元)','直接下单ROI', '7日间接成交金额(元)', '播放数', '3s播放数', '播放完成数','广告类型', '达人uid', '成交智能优惠券金额','订单成交成本']
        df = pd.DataFrame(columns=table_title)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        df.to_excel(filepath, index=False, sheet_name='Sheet1')
        print(f'文件已保存至 {filepath}')

#写入计划数据
def write_data(datafile,yesterday):
    dir = os.path.dirname(__file__)
    filepath = os.path.join(dir,'源表','推广费',f'推广费-{yesterday}.xlsx')
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb.active
    start_row = ws.max_row + 1
    for idx, ad_detail in enumerate(datafile, start=start_row):
        for col_num, value in enumerate(ad_detail, start=1):
            ws.cell(row=idx, column=col_num).value = value
    wb.save(filepath)
    print(f'数据保存至路径：{filepath}')

#获取店铺id
def get_shop_id(access_token,retry_limit=3):
    open_api_url = 'https://ad.oceanengine.com/open_api/'
    url_params = 'oauth2/advertiser/get/'
    url = open_api_url + url_params
    json_params = {
        'access_token': access_token
    }
    try:
        retry_count = 1
        while retry_count <= retry_limit:
            format_url = build_url(url,json_params)
            rsp = get(format_url)
            rsp_data = rsp.json()
            data_list = rsp_data['data'].get('list',[])
            if data_list:
                shop_id_list = [item['advertiser_id'] for item in data_list if item['account_role'] == 'CUSTOMER_OPERATOR' and item['is_valid']]
                return shop_id_list
            else:
                retry_count += 1
                sleep(2)
        else:
            print(f'get_shop_id请求多次失败: {rsp_data}')
            return None
    except Exception as e:
        print(f'get_shop_id未成功发起请求: {e}')
        return None

#获取账户id列表
def get_adv_info_from_flow(yesterday):
    dir = os.path.dirname(__file__)
    filepath = os.path.join(dir, '财务流水')
    adv = pd.read_excel(fr'{filepath}\财务流水-{yesterday}.xlsx',dtype={'账户ID':str})
    result_df = adv.groupby(['账户ID', '账户名称'])[['消返红包消耗', '余额总消耗', '共享赠款消耗', '共享钱包消耗']].sum()
    result_df['总消耗'] = result_df.sum(axis=1)
    result_df = result_df.reset_index()
    result_df = result_df[['账户ID', '账户名称', '总消耗']]
    result = result_df[result_df['总消耗'] > 0]
    all_advertiser_id = result['账户ID'].tolist()
    all_advertiser_name = result['账户名称'].tolist()
    #将两个列表改为id与name对应的列表形式：
    #combined_list = [[x, y] for x, y in zip(list1, list2)]
    advertiser_list = [[x, y] for x, y in zip_longest(all_advertiser_id, all_advertiser_name, fillvalue='')]
    return advertiser_list

#获取标准商品账户id列表
def get_adv_info_from_product(yesterday):
    dir = os.path.dirname(__file__)
    filepath = os.path.join(dir, '源表','推广费')
    adv = pd.read_excel(fr'{filepath}\推广费-{yesterday}.xlsx',dtype='str')
    adv = adv[adv['推广类型'] == '标准推商品']
    all_advertiser_id = adv['账户ID'].tolist()
    all_advertiser_name = adv['账户名称'].tolist()
    all_plan_id = adv['计划ID'].tolist()
    all_plan_name = adv['计划名称'].tolist()
    #将两个列表改为id与name对应的列表形式：
    #combined_list = [[x, y] for x, y in zip(list1, list2)]
    advertiser_list = [[x, y, z, w] for x, y, z, w in zip_longest(all_advertiser_id, all_advertiser_name, all_plan_id, all_plan_name, fillvalue='')]
    return advertiser_list

#获取标准直播账户id列表
def get_adv_info_from_live(yesterday):
    dir = os.path.dirname(__file__)
    filepath = os.path.join(dir, '源表','推广费')
    adv = pd.read_excel(fr'{filepath}\推广费-{yesterday}.xlsx',dtype='str')
    adv = adv[adv['推广类型'] == '标准推直播']
    if adv.empty:
        return None
    all_advertiser_id = adv['账户ID'].tolist()
    all_advertiser_name = adv['账户名称'].tolist()
    all_plan_id = adv['计划ID'].tolist()
    all_plan_name = adv['计划名称'].tolist()
    #将两个列表改为id与name对应的列表形式：
    #combined_list = [[x, y] for x, y in zip(list1, list2)]
    advertiser_list = [[x, y, z, w] for x, y, z, w in zip_longest(all_advertiser_id, all_advertiser_name, all_plan_id, all_plan_name, fillvalue='')]
    return advertiser_list

def get_aweme_info_from_live(yesterday):
    dir = os.path.dirname(__file__)
    filepath = os.path.join(dir, '源表','推广费')
    adv = pd.read_excel(fr'{filepath}\推广费-{yesterday}.xlsx',dtype='str')
    adv = adv[adv['推广类型'] == '标准推直播']
    if adv.empty:
        return None
    all_advertiser_id = adv['账户ID'].tolist()
    all_advertiser_name = adv['账户名称'].tolist()
    all_plan_id = adv['计划ID'].tolist()
    all_plan_name = adv['计划名称'].tolist()
    all_aweme_uid = adv['达人uid'].tolist()
    #将两个列表改为id与name对应的列表形式：
    #combined_list = [[x, y] for x, y in zip(list1, list2)]
    advertiser_list = [[x, y, z, w , v] for x, y, z, w, v in zip_longest(all_advertiser_id, all_advertiser_name, all_plan_id, all_plan_name, all_aweme_uid, fillvalue='')]
    return advertiser_list

#获取全域商品账户id列表
def get_adv_info_from_uni_product(yesterday):
    dir = os.path.dirname(__file__)
    filepath = os.path.join(dir, '源表','推广费')
    adv = pd.read_excel(fr'{filepath}\推广费-{yesterday}.xlsx',dtype='str')
    adv = adv[adv['推广类型'] == '全域推商品']
    all_advertiser_id = adv['账户ID'].tolist()
    all_advertiser_name = adv['账户名称'].tolist()
    all_plan_id = adv['计划ID'].tolist()
    all_plan_name = adv['计划名称'].tolist()
    #将两个列表改为id与name对应的列表形式：
    #combined_list = [[x, y] for x, y in zip(list1, list2)]
    advertiser_list = [[x, y, z, w] for x, y, z, w in zip_longest(all_advertiser_id, all_advertiser_name, all_plan_id, all_plan_name, fillvalue='')]
    return advertiser_list

#获取全域直播账户id列表
def get_adv_info_from_uni_live(yesterday):
    dir = os.path.dirname(__file__)
    filepath = os.path.join(dir, '源表','推广费')
    adv = pd.read_excel(fr'{filepath}\推广费-{yesterday}.xlsx',dtype='str')
    adv = adv[adv['推广类型'] == '全域推直播']
    all_advertiser_id = adv['账户ID'].tolist()
    all_advertiser_name = adv['账户名称'].tolist()
    all_aweme_uid = adv['达人uid'].tolist()
    all_aweme_name = adv['抖音号名称'].tolist()
    #将两个列表改为id与name对应的列表形式：
    #combined_list = [[x, y] for x, y in zip(list1, list2)]
    advertiser_list = [[x, y, z, w] for x, y, z, w in zip_longest(all_advertiser_id, all_advertiser_name, all_aweme_uid, all_aweme_name, fillvalue='')]
    return advertiser_list

#写入素材数据
def write_material_data(datafile,yesterday):
    dir = os.path.dirname(__file__)
    filepath = os.path.join(dir,'源表','素材报表',f'素材报表-{yesterday}.xlsx')
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb.active
    start_row = ws.max_row + 1
    for idx, ad_detail in enumerate(datafile, start=start_row):
        for col_num, value in enumerate(ad_detail, start=1):
            ws.cell(row=idx, column=col_num).value = value
    wb.save(filepath)
    print(f'数据保存至路径：{filepath}')